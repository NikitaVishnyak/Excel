from PyQt5 import QtCore, QtGui, QtWidgets
from string import ascii_uppercase
from openpyxl import load_workbook
import pandas as pd
import itertools
import os

from cell import Cell
from parser import Parser




class mainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self.tableWidget = None
        self.msg = MessageBox()
        self.external_table = XlsxData()
        self.cell = None
        self.textInInputLine = ''
        self.rowCount = 6
        self.colCount = 6
        self.setObjectName("Excel")
        self.setMinimumSize(860, 500)
        self.setMaximumSize(860, 500)
        self.init_ui()
        self.on_event()
        self.clearTable()
        self.is_saved = True

    def init_ui(self):
        self.centralWidget = QtWidgets.QWidget(self)
        self.centralWidget.setObjectName("centralWidget")

        self.addRowButton = QtWidgets.QPushButton(self.centralWidget)
        self.addRowButton.setGeometry(QtCore.QRect(46, 10, 110, 30))
        self.addRowButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.addRowButton.setObjectName("addRowButton")

        self.addColButton = QtWidgets.QPushButton(self.centralWidget)
        self.addColButton.setGeometry(QtCore.QRect(174, 10, 110, 30))
        self.addColButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.addColButton.setObjectName("addColButton")

        self.delRowButton = QtWidgets.QPushButton(self.centralWidget)
        self.delRowButton.setGeometry(QtCore.QRect(46, 50, 110, 30))
        self.delRowButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.delRowButton.setObjectName("delRowButton")

        self.delColButton = QtWidgets.QPushButton(self.centralWidget)
        self.delColButton.setGeometry(QtCore.QRect(174, 50, 110, 30))
        self.delColButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.delColButton.setObjectName("delColButton")

        self.calculateButton = QtWidgets.QPushButton(self.centralWidget)
        self.calculateButton.setGeometry(QtCore.QRect(760, 30, 50, 30))
        self.calculateButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.calculateButton.setObjectName("calculateButton")

        self.clearButton = QtWidgets.QPushButton(self.centralWidget)
        self.clearButton.setGeometry(QtCore.QRect(301, 30, 50, 30))
        self.clearButton.setCursor(QtGui.QCursor(QtCore.Qt.PointingHandCursor))
        self.clearButton.setObjectName("clearButton")

        self.lineEdit = QtWidgets.QLineEdit(self.centralWidget)
        self.lineEdit.setGeometry(QtCore.QRect(580, 30, 150, 30))
        self.lineEdit.setObjectName("lineEdit")

        self.tableWidget = QtWidgets.QTableWidget(self.centralWidget)
        self.tableWidget.setGeometry(QtCore.QRect(20, 90, 821, 351))
        self.tableWidget.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.tableWidget.setLineWidth(1)
        self.tableWidget.setMidLineWidth(0)
        self.tableWidget.setShowGrid(True)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(6)
        self.tableWidget.setRowCount(6)

        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setVerticalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setVerticalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setVerticalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setVerticalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setVerticalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setVerticalHeaderItem(5, item)

        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        self.tableWidget.setHorizontalHeaderItem(5, item)

        self.setCentralWidget(self.centralWidget)

        self.menubar = QtWidgets.QMenuBar(self)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1080, 720))
        self.menubar.setObjectName("menubar")

        self.menuFile = QtWidgets.QMenu(self.menubar)
        self.menuFile.setObjectName("menuFile")

        self.actionOpen = QtWidgets.QAction(self)
        self.actionOpen.setObjectName("actionOpen")
        self.menuFile.addAction(self.actionOpen)

        self.actionSave = QtWidgets.QAction(self)
        self.actionSave.setObjectName("actionSave")
        self.menuFile.addAction(self.actionSave)

        self.menubar.addAction(self.menuFile.menuAction())
        self.setMenuBar(self.menubar)

        self.retranslate_ui(self)
        QtCore.QMetaObject.connectSlotsByName(self)

    def retranslate_ui(self, excel):
        _translate = QtCore.QCoreApplication.translate
        excel.setWindowTitle(_translate("Excel", "Excel"))

        self.addRowButton.setText(_translate("Excel", "ADD ROW"))
        self.addColButton.setText(_translate("Excel", "ADD COLUMN"))
        self.delRowButton.setText(_translate("Excel", "DELETE ROW"))
        self.delColButton.setText(_translate("Excel", "DELETE COLUMN"))
        self.calculateButton.setText(_translate("Excel", "="))
        self.clearButton.setText(_translate("Excel", "C"))

        self.tableWidget.setSortingEnabled(False)

        item = self.tableWidget.verticalHeaderItem(0)
        item.setText(_translate("Excel", "1"))
        item = self.tableWidget.verticalHeaderItem(1)
        item.setText(_translate("Excel", "2"))
        item = self.tableWidget.verticalHeaderItem(2)
        item.setText(_translate("Excel", "3"))
        item = self.tableWidget.verticalHeaderItem(3)
        item.setText(_translate("Excel", "4"))
        item = self.tableWidget.verticalHeaderItem(4)
        item.setText(_translate("Excel", "5"))
        item = self.tableWidget.verticalHeaderItem(5)
        item.setText(_translate("Excel", "6"))

        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("Excel", "A"))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("Excel", "B"))
        item = self.tableWidget.horizontalHeaderItem(2)
        item.setText(_translate("Excel", "C"))
        item = self.tableWidget.horizontalHeaderItem(3)
        item.setText(_translate("Excel", "D"))
        item = self.tableWidget.horizontalHeaderItem(4)
        item.setText(_translate("Excel", "E"))
        item = self.tableWidget.horizontalHeaderItem(5)
        item.setText(_translate("Excel", "F"))

        self.menuFile.setTitle(_translate("Excel", "File"))
        self.actionOpen.setText(_translate("Excel", "Open"))
        self.actionSave.setText(_translate("Excel", "Save"))


        __sortingEnabled = self.tableWidget.isSortingEnabled()
        self.tableWidget.setSortingEnabled(False)
        self.tableWidget.setSortingEnabled(__sortingEnabled)

    def on_event(self):
        self.addRowButton.clicked.connect(self.addRow)
        self.delRowButton.clicked.connect(self.delRow)
        self.addColButton.clicked.connect(self.addCol)
        self.delColButton.clicked.connect(self.delCol)
        self.calculateButton.clicked.connect(self.calculate)
        self.clearButton.clicked.connect(self.clearLine)
        self.actionOpen.triggered.connect(self.openDoc)
        self.actionSave.triggered.connect(self.saveDoc)

        self.tableWidget.selectionModel().selectionChanged.connect(self.getSelectedSell)
        self.tableWidget.itemChanged.connect(self.trackChanges)

    def getSelectedSell(self, selected, deselected):
        for ix in selected.indexes():
            self.cell = Cell(ix.row(), ix.column(), self.tableWidget)

    def closingApp(self, event):
        if self.is_saved != True:
            self.msg.closingFile(event, self.tableWidget, self.saveDoc)

    def addRow(self):
        self.tableWidget.setRowCount(self.rowCount + 1)

        self.tableWidget.setVerticalHeaderItem(self.rowCount, QtWidgets.QTableWidgetItem(str(self.rowCount + 1)))

        self.rowCount += 1

    def addCol(self):
        base_char = 1
        for i in iterAllStrings():
            if base_char > self.colCount:
                self.tableWidget.setColumnCount(self.colCount + 1)
                self.tableWidget.setHorizontalHeaderItem(self.colCount, QtWidgets.QTableWidgetItem(i))

                base_char += 1
                self.colCount += 1
                break
            else:
                base_char += 1
                continue

    def delRow(self):
        if self.rowCount > 1:
            self.tableWidget.setRowCount(self.rowCount - 1)
            self.rowCount -= 1
        else:
            self.msg.minimalAmountOfRow()

    def delCol(self):
        if self.colCount > 1:
            self.tableWidget.setColumnCount(self.colCount - 1)
            self.colCount -= 1
        else:
            self.msg.minimalAmountOfCol()

    def calculate(self):
        expression = self.lineEdit.text()
        if self.cell:
            if expression != '':
                self.cell.parsing(expression)
            else:
                self.msg.noExpressionInCell()
        else:
            self.msg.selectCellWarning()

    def openDoc(self):
        if self.is_saved:
            self.fillTable()
        else:
            if self.msg.reopeningFile(self.tableWidget, self.saveDoc):
                self.fillTable()

    def saveDoc(self):
        try:
            path = self.msg.savingFile(self)

            columnHeaders = []
            for k in range(self.tableWidget.model().columnCount()):
                columnHeaders.append(self.tableWidget.horizontalHeaderItem(k).text())

            dataFrame = pd.DataFrame(columns = columnHeaders)
            for row in range(self.tableWidget.rowCount()):
                for col in range(self.tableWidget.columnCount()):
                    dataFrame.at[row, columnHeaders[col]] = self.tableWidget.item(row, col).text()

            dataFrame.to_excel(path[0], header = False, index = False)

            self.is_saved = True
        except:
            self.msg.wrongFileType()

    def fillTable(self):
        try:
            self.external_table.setPath(self.msg.openingFile(self))
            self.external_table.reloadWorkbook()
            self.clearTable()

            table_data = list(self.external_table.getWorksheet().values)
            self.blankForData()

            row_ix = 0
            for value_tuple in table_data:
                col_ix = 0
                for value in value_tuple:
                    if value is not None:
                        self.tableWidget.setItem(row_ix, col_ix, QtWidgets.QTableWidgetItem(str(value)))
                    else:
                        self.tableWidget.setItem(row_ix, col_ix, QtWidgets.QTableWidgetItem(''))
                    col_ix += 1
                row_ix += 1

            self.rowCount = self.tableWidget.rowCount()
            self.colCount = self.tableWidget.columnCount()
            self.is_saved = True
        except:
            self.msg.wrongFileType()

    def clearTable(self):
        for row in range(0, self.rowCount):
            for col in range(0, self.colCount):
                self.tableWidget.setItem(row, col, QtWidgets.QTableWidgetItem())
                col += 1
            row += 1


    def blankForData(self):
        maxRow = self.external_table.getWorksheet().max_row
        maxCol = self.external_table.getWorksheet().max_column
        for it in range(3, maxRow):
            if self.rowCount < maxRow:
                self.addRow()
        for it in range(3, maxCol):
            if self.colCount < maxCol:
                self.addCol()

    def clearLine(self):
        self.lineEdit.setText('')

    def addMoreText(self, txt):
        expression = self.lineEdit.text()
        self.textInInputLine = expression + txt
        return self.textInInputLine

    def trackChanges(self):
        self.is_saved = False
        row = self.tableWidget.currentIndex().row()
        col = self.tableWidget.currentIndex().column()
        element = self.tableWidget.item(row, col)
        if element is not None and element.text() != '':
            cell = Cell(row, col, self.tableWidget)
            if cell.getExpressionFromCell(row, col)[0] == '=' or cell.getExpressionFromCell(row, col)[0] == '#':
                cell.parsing(element.text())

class XlsxData:
    def __init__(self):
        self.path = None

    def getWorksheet(self):
        return self.ws

    def setPath(self, path):
        self.path = path[0]

    def reloadWorkbook(self):
        self.wb = load_workbook(self.path)
        self.ws = self.wb.active


class MessageBox:
    def __init__(self):
        self.msg = QtWidgets.QMessageBox()

    def minimalAmountOfRow(self):
        self.msg.setIcon(self.msg.Warning)
        self.msg.setWindowTitle("Warning!")
        self.msg.setText("A table cannot contain less than 1 row")
        self.msg.setStandardButtons(self.msg.Ok)
        self.msg.exec_()

    def minimalAmountOfCol(self):
        self.msg.setIcon(self.msg.Warning)
        self.msg.setWindowTitle("Warning!")
        self.msg.setText("A table cannot contain less than 1 column")
        self.msg.setStandardButtons(self.msg.Ok)
        self.msg.exec_()


    def wrongEntry(self):
        self.msg.setIcon(self.msg.Warning)
        self.msg.setWindowTitle("Warning!")
        self.msg.setText("Wrong entry!")
        self.msg.setStandardButtons(self.msg.Ok)
        self.msg.exec_()

    def selectCellWarning(self):
        self.msg.setIcon(self.msg.Warning)
        self.msg.setWindowTitle("Warning!")
        self.msg.setText("You have not selected the cell")
        self.msg.setStandardButtons(self.msg.Ok)
        self.msg.exec_()

    def noExpressionInCell(self):
        self.msg.setIcon(self.msg.Warning)
        self.msg.setWindowTitle("Warning!")
        self.msg.setText("You have not written expression")
        self.msg.setStandardButtons(self.msg.Ok)
        self.msg.exec_()

    def incorrectExpression(self):
        self.msg.setIcon(self.msg.Warning)
        self.msg.setWindowTitle("Warning!")
        self.msg.setText("Incorrect expression")
        self.msg.setStandardButtons(self.msg.Ok)
        self.msg.exec_()

    def closingFile(self, event, widget, saving_func):
        reply = self.msg.question(widget, 'Closing', 'Save changes?', self.msg.No | self.msg.Yes)
        if reply == self.msg.No:
            event.accept()
        elif reply == self.msg.Yes:
            saving_func()
            event.ignore()

    def reopeningFile(self, widget, saving_func):
        reply = self.msg.question(widget, 'Closing', 'Open new file without saving previous?', self.msg.Yes | self.msg.Save)
        if reply == self.msg.Save:
            saving_func()
        elif reply == self.msg.Yes:
            return True

    def openingFile(self, parent):
        return QtWidgets.QFileDialog.getOpenFileName(parent, 'Select a file', os.getcwd(), 'Excel File (*.xlsx *.xls)', 'Excel File (*.xlsx *.xls)')

    def savingFile(self, parent):
        return QtWidgets.QFileDialog.getSaveFileName(parent, 'Select a file', '', 'Excel File (*.xlsx *.xls)', 'Excel File (*.xlsx *.xls)')

    def wrongFileType(self):
        self.msg.setIcon(self.msg.Warning)
        self.msg.setWindowTitle("Warning!")
        self.msg.setText("Wrong file format")
        self.msg.setStandardButtons(self.msg.Ok)
        self.msg.exec_()

    def zeroDeviding(self):
        self.msg.setIcon(self.msg.Warning)
        self.msg.setWindowTitle("Warning!")
        self.msg.setText("You can't divide by zero")
        self.msg.setStandardButtons(self.msg.Ok)
        self.msg.exec_()

def iterAllStrings():
    for size in itertools.count(1):
        for s in itertools.product(ascii_uppercase, repeat=size):
            yield "".join(s)